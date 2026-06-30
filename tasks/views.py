from rest_framework import viewsets, permissions, status
from rest_framework.response import Response
from rest_framework.decorators import action
from django.shortcuts import get_object_or_404
from django.db import transaction
from django.db.models import Q

from .models import Task, TaskComment, ActivityLog, Notification, EmailLog
from .serializers import (
    TaskSerializer, TaskMinSerializer, TaskCommentSerializer,
    ActivityLogSerializer, NotificationSerializer, EmailLogSerializer
)
from tables.permissions import has_table_access

class TaskViewSet(viewsets.ModelViewSet):
    serializer_class = TaskSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        # Users see tasks belonging to accessible tables
        user = self.request.user
        if user.role == "SUPER_ADMIN":
            return Task.objects.all()

        # Get rows in user's accessible tables
        from tables.permissions import get_accessible_tables
        accessible_tables = get_accessible_tables(user)
        return Task.objects.filter(row__table__in=accessible_tables)

    @action(detail=True, methods=["post"], url_path="update-status")
    @transaction.atomic
    def update_status(self, request, pk=None):
        task = get_object_or_404(Task, pk=pk)
        new_status = request.data.get("status")

        if new_status not in dict(Task.STATUS_CHOICES):
            return Response({"error": "Invalid status value"}, status=status.HTTP_400_BAD_REQUEST)

        # Permission check: must have edit access to the table OR be assigned to the task and updating status to COMPLETED
        is_assigned = task.assigned_to.filter(id=request.user.id).exists()
        if not (has_table_access(request.user, task.row.table, "EDIT") or (new_status == "COMPLETED" and is_assigned)):
            return Response({"error": "No permission to update this task status"}, status=status.HTTP_403_FORBIDDEN)

        old_status = task.status
        task.status = new_status
        task.save(update_fields=["status"])

        # Sync back to STATUS cell if such a column exists
        from tables.models import Column, CellValue
        status_col = Column.objects.filter(table=task.row.table, name__iexact="STATUS").first()
        if status_col:
            CellValue.objects.update_or_create(
                row=task.row, column=status_col,
                defaults={"value": new_status, "updated_by": request.user}
            )

        # Create Activity Log
        ActivityLog.objects.create(
            task=task,
            action=f"Changed status from {old_status} to {new_status}",
            user=request.user,
            details={"old_status": old_status, "new_status": new_status}
        )

        # Trigger notification to admin if employee updates task to "Ready for Review"
        if new_status == "READY_FOR_REVIEW":
            # Find department admin / admin / super admin
            admins = []
            if task.row.table.department:
                from auth_app.models import EmployeeUser
                admins = EmployeeUser.objects.filter(
                    role__in=["ADMIN", "DEPARTMENT_ADMIN"],
                    department=task.row.table.department
                )
            if not admins:
                from auth_app.models import EmployeeUser
                admins = EmployeeUser.objects.filter(role="SUPER_ADMIN")

            for admin in admins:
                Notification.objects.create(
                    user=admin,
                    task=task,
                    title="Task Ready for Review",
                    description=f"Task '{task.row.cells.filter(column__name='TASK_NAME').first().value}' has been marked ready for review by {request.user.full_name}",
                    type="REVIEW"
                )

        # If admin approves/rejects/requests changes:
        if old_status == "READY_FOR_REVIEW" and request.user.role in ["SUPER_ADMIN", "ADMIN", "DEPARTMENT_ADMIN"]:
            # Notify employees assigned to the task
            for employee in task.assigned_to.all():
                Notification.objects.create(
                    user=employee,
                    task=task,
                    title=f"Task Status Updated: {new_status}",
                    description=f"Admin {request.user.full_name} updated task status to {new_status}.",
                    type="SYSTEM"
                )

        return Response(TaskSerializer(task).data, status=status.HTTP_200_OK)

    @action(detail=True, methods=["post"], url_path="assign")
    @transaction.atomic
    def assign_task(self, request, pk=None):
        task = get_object_or_404(Task, pk=pk)
        
        # Only admins or table owners can assign tasks
        if not has_table_access(request.user, task.row.table, "ADMIN"):
            return Response({"error": "Only admins can assign tasks"}, status=status.HTTP_403_FORBIDDEN)

        assigned_to_ids = request.data.get("assigned_to", [])
        from auth_app.models import EmployeeUser
        employees = EmployeeUser.objects.filter(id__in=assigned_to_ids)
        
        # Track old assignees
        old_assignees = list(task.assigned_to.values_list("email", flat=True))
        task.assigned_to.set(employees)
        task.assigned_by = request.user
        task.save(update_fields=["assigned_by"])

        new_assignees = list(employees.values_list("email", flat=True))

        # Log action
        ActivityLog.objects.create(
            task=task,
            action="Assigned Task",
            user=request.user,
            details={"old_assignees": old_assignees, "new_assignees": new_assignees}
        )

        # Notify employees
        for employee in employees:
            Notification.objects.create(
                user=employee,
                task=task,
                title="Task Assigned to You",
                description=f"You have been assigned a task by {request.user.full_name}",
                type="ASSIGNED"
            )

        return Response(TaskSerializer(task).data, status=status.HTTP_200_OK)

    @action(detail=False, methods=["get"], url_path="recipients")
    def list_recipients(self, request):
        from auth_app.models import EmployeeUser
        users = EmployeeUser.objects.filter(is_active=True).order_by("full_name", "email")
        data = [{"id": u.id, "full_name": u.full_name, "email": u.email, "role": u.role} for u in users]
        return Response(data, status=status.HTTP_200_OK)

    @action(detail=True, methods=["post"], url_path="flag-issue")
    @transaction.atomic
    def flag_issue(self, request, pk=None):
        task = get_object_or_404(Task, pk=pk)
        recipient_id = request.data.get("recipient_id")
        description = request.data.get("description")

        if not recipient_id or not description:
            return Response({"error": "recipient_id and description are required"}, status=status.HTTP_400_BAD_REQUEST)

        from auth_app.models import EmployeeUser
        recipient = get_object_or_404(EmployeeUser, id=recipient_id, is_active=True)

        # 1. Create task comment record of the issue
        comment = TaskComment.objects.create(
            task=task,
            author=request.user,
            content=f"[Issue Flagged to {recipient.full_name or recipient.email}] Description: {description}"
        )

        # 2. Create notification for recipient
        Notification.objects.create(
            user=recipient,
            task=task,
            title="Task Issue Flagged",
            description=f"Issue reported by {request.user.full_name or request.user.email}: {description}",
            type="COMMENT"
        )

        # Create Activity Log
        ActivityLog.objects.create(
            task=task,
            action="Flagged Issue",
            user=request.user,
            details={"recipient_id": recipient.id, "comment_id": comment.id}
        )

        return Response(TaskSerializer(task).data, status=status.HTTP_200_OK)

    @action(detail=True, methods=["post"], url_path="log-follow-up")
    @transaction.atomic
    def log_follow_up(self, request, pk=None):
        task = get_object_or_404(Task, pk=pk)
        
        # Verify table is SALES or LIST_PID
        if task.row.table.job_type not in ["SALES", "LIST_PID"]:
            return Response({"error": "This action is only supported for Sales or LIST_PID tasks."}, status=status.HTTP_400_BAD_REQUEST)
        
        discussed_points = request.data.get("discussed_points")
        new_status = request.data.get("status")
        next_follow_up_date_str = request.data.get("next_follow_up_date")
        
        if not discussed_points or not discussed_points.strip():
            return Response({"error": "Discussed points are required."}, status=status.HTTP_400_BAD_REQUEST)
            
        if not new_status:
            return Response({"error": "Status is required."}, status=status.HTTP_400_BAD_REQUEST)
            
        if new_status not in dict(Task.STATUS_CHOICES):
            return Response({"error": "Invalid status value."}, status=status.HTTP_400_BAD_REQUEST)
        
        # If status is continuing (PENDING, IN_PROGRESS, READY_FOR_REVIEW), next follow-up date is mandatory
        is_continuing = new_status in ["PENDING", "IN_PROGRESS", "READY_FOR_REVIEW"]
        next_follow_up_date = None
        
        if is_continuing:
            if not next_follow_up_date_str:
                return Response({"error": "Next follow-up date is required when task is in progress."}, status=status.HTTP_400_BAD_REQUEST)
            try:
                from datetime import datetime
                next_follow_up_date = datetime.strptime(next_follow_up_date_str.split("T")[0], "%Y-%m-%d").date()
            except ValueError:
                return Response({"error": "Invalid date format for next follow-up date. Use YYYY-MM-DD."}, status=status.HTTP_400_BAD_REQUEST)
        
        # Save TaskFollowUp record
        # Current follow-up date is task.due_date
        current_follow_up_date = task.due_date
        
        from .models import TaskFollowUp
        follow_up = TaskFollowUp.objects.create(
            task=task,
            follow_up_date=current_follow_up_date,
            discussed_points=discussed_points,
            next_follow_up_date=next_follow_up_date,
            entered_by=request.user
        )
        
        # Update Task Status and due_date
        old_status = task.status
        task.status = new_status
        old_due_date = task.due_date
        
        if is_continuing and next_follow_up_date:
            task.due_date = next_follow_up_date
            task.alert_mail_sent = False  # Reset alert_mail_sent so alert is sent on the new date
        task.save()
        
        # Sync ALERT_MAIL state back to the spreadsheet row to ensure cell is NO
        from .tasks import update_task_row_mail_columns
        update_task_row_mail_columns(task)
        
        # Sync back to STATUS cell if such a column exists
        from tables.models import Column, CellValue
        status_col = Column.objects.filter(table=task.row.table, name__iexact="STATUS").first()
        if status_col:
            CellValue.objects.update_or_create(
                row=task.row, column=status_col,
                defaults={"value": new_status, "updated_by": request.user}
            )
            
        # Sync next follow up date to FOLLOW_UP_DATE column if continuing
        follow_up_col = Column.objects.filter(table=task.row.table, name__iexact="FOLLOW_UP_DATE").first()
        if follow_up_col and is_continuing and next_follow_up_date:
            CellValue.objects.update_or_create(
                row=task.row, column=follow_up_col,
                defaults={"value": next_follow_up_date.isoformat(), "updated_by": request.user}
            )
            
        # Sync next follow up date to DUE_DATE_FLOW_FORCE column if continuing
        flow_force_col = Column.objects.filter(table=task.row.table, name__iexact="DUE_DATE_FLOW_FORCE").first()
        if flow_force_col and is_continuing and next_follow_up_date:
            CellValue.objects.update_or_create(
                row=task.row, column=flow_force_col,
                defaults={"value": next_follow_up_date.isoformat(), "updated_by": request.user}
            )
            
        # Sync next follow up date to DUE_DATE column just in case (as fall back)
        due_date_col = Column.objects.filter(table=task.row.table, name__iexact="DUE_DATE").first()
        if due_date_col and is_continuing and next_follow_up_date:
            CellValue.objects.update_or_create(
                row=task.row, column=due_date_col,
                defaults={"value": next_follow_up_date.isoformat(), "updated_by": request.user}
            )
            
        # Create TaskComment of the discussion points as internal reference / chat history
        TaskComment.objects.create(
            task=task,
            author=request.user,
            content=f"[Logged Follow-up] Discussion: {discussed_points}" + (f"\nNext Follow-up scheduled for: {next_follow_up_date.isoformat()}" if next_follow_up_date else "\nNo further follow-up required (closed/cancelled).")
        )

        # Create additional comment for sales follow-up
        old_date_str = current_follow_up_date.strftime("%Y-%m-%d") if current_follow_up_date else "N/A"
        TaskComment.objects.create(
            task=task,
            author=request.user,
            content=f"enter new follow up under the old follow up date: {old_date_str}"
        )
        
        # Create Activity Log
        ActivityLog.objects.create(
            task=task,
            action="Logged Follow-up",
            user=request.user,
            details={
                "follow_up_date": current_follow_up_date.isoformat() if current_follow_up_date else None,
                "next_follow_up_date": next_follow_up_date.isoformat() if next_follow_up_date else None,
                "discussed_points": discussed_points,
                "new_status": new_status
            }
        )
        
        # Trigger notification to all assigned users
        if is_continuing and next_follow_up_date:
            for assignee in task.assigned_to.all():
                Notification.objects.create(
                    user=assignee,
                    task=task,
                    title="Next Follow-up Scheduled",
                    description=f"A new follow-up for Customer '{task.task_name}' has been scheduled for {next_follow_up_date.isoformat()} by {request.user.full_name}",
                    type="SYSTEM"
                )
                
        return Response(TaskSerializer(task).data, status=status.HTTP_200_OK)

class TaskCommentViewSet(viewsets.ModelViewSet):
    serializer_class = TaskCommentSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        task_id = self.request.query_params.get("task")
        if not task_id:
            return TaskComment.objects.none()
        return TaskComment.objects.filter(task_id=task_id)

    def perform_create(self, serializer):
        task = serializer.validated_data["task"]
        # Save comment and log it
        comment = serializer.save(author=self.request.user)

        ActivityLog.objects.create(
            task=task,
            action="Added Comment",
            user=self.request.user,
            details={"comment_id": comment.id, "is_internal_note": comment.is_internal_note}
        )

        # Notification for comment mentions or alerts
        recipients = set()
        if task.assigned_by and task.assigned_by.id != self.request.user.id:
            recipients.add(task.assigned_by)
        for user in task.assigned_to.exclude(id=self.request.user.id):
            recipients.add(user)
        table_creator = task.row.table.created_by
        if table_creator and table_creator.id != self.request.user.id:
            recipients.add(table_creator)

        for recipient in recipients:
            Notification.objects.create(
                user=recipient,
                task=task,
                title="New Comment on Task",
                description=f"{self.request.user.full_name} commented on your task.",
                type="COMMENT"
            )

class ActivityLogViewSet(viewsets.ReadOnlyModelViewSet):
    serializer_class = ActivityLogSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        task_id = self.request.query_params.get("task")
        if not task_id:
            return ActivityLog.objects.none()
        return ActivityLog.objects.filter(task_id=task_id)

class NotificationViewSet(viewsets.ModelViewSet):
    serializer_class = NotificationSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        return Notification.objects.filter(user=self.request.user)

    @action(detail=True, methods=["post"], url_path="read")
    def mark_read(self, request, pk=None):
        notification = self.get_object_or_404(pk)
        notification.is_read = True
        notification.save(update_fields=["is_read"])
        return Response({"status": "read"}, status=status.HTTP_200_OK)

    @action(detail=False, methods=["post"], url_path="read-all")
    def mark_all_read(self, request):
        Notification.objects.filter(user=request.user, is_read=False).update(is_read=True)
        return Response({"status": "all read"}, status=status.HTTP_200_OK)

    def get_object_or_404(self, pk):
        return get_object_or_404(Notification, pk=pk, user=self.request.user)

from django.contrib.auth.decorators import login_required
from django.shortcuts import redirect

@login_required
def notification_mark_read_view(request, notification_id):
    # Try tasks app Notification first
    from tasks.models import Notification as TasksNotification
    try:
        notification = TasksNotification.objects.get(id=notification_id, user=request.user)
        notification.is_read = True
        notification.save(update_fields=["is_read"])
        if notification.task:
            from tables.permissions import has_table_access
            is_admin_or_super = request.user.role in ["SUPER_ADMIN", "ADMIN"]
            has_access = has_table_access(request.user, notification.task.row.table, "VIEW")
            if is_admin_or_super or has_access:
                return redirect(f"/tables/{notification.task.row.table_id}/?open_task_id={notification.task.id}")
            else:
                return redirect(f"/tasks/{notification.task.id}/detail/")
        return redirect("/")
    except TasksNotification.DoesNotExist:
        # Fallback to task_tracker app Notification
        from task_tracker.models import Notification as TrackerNotification
        try:
            notification = TrackerNotification.objects.get(id=notification_id, user=request.user)
            notification.read = True
            notification.save(update_fields=["read"])
            if notification.row:
                return redirect("task_tracker:task_detail", tracker_id=notification.row.tracker.id, task_id=notification.row.id)
            return redirect("task_tracker:tracker_dashboard")
        except TrackerNotification.DoesNotExist:
            pass
    return redirect("/")

from django.shortcuts import render
import csv
import openpyxl
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table as RLTable, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from django.http import HttpResponse
from tables.models import CellValue

@login_required
def reports_view(request):
    # Only Admin, Department Admin, and Super Admin can see report generator
    if request.user.role not in ["SUPER_ADMIN", "ADMIN", "DEPARTMENT_ADMIN"]:
        return redirect("/")

    from auth_app.models import EmployeeUser
    # Filter users based on department
    if request.user.role in ["ADMIN", "DEPARTMENT_ADMIN"] and request.user.department:
        employees = EmployeeUser.objects.filter(department=request.user.department)
    else:
        employees = EmployeeUser.objects.all()

    # Query tasks
    if request.user.role == "SUPER_ADMIN":
        tasks = Task.objects.all()
    elif request.user.role in ["ADMIN", "DEPARTMENT_ADMIN"]:
        if request.user.department:
            tasks = Task.objects.filter(row__table__department=request.user.department)
        else:
            tasks = Task.objects.all()

    # Apply filters
    employee_id = request.GET.get("employee")
    month = request.GET.get("month")
    date_from = request.GET.get("date_from")
    date_to = request.GET.get("date_to")

    if employee_id:
        tasks = tasks.filter(assigned_to__id=employee_id)
    if month:
        # Format: YYYY-MM
        try:
            year_val, month_val = map(int, month.split("-"))
            tasks = tasks.filter(due_date__year=year_val, due_date__month=month_val)
        except ValueError:
            pass
    if date_from:
        tasks = tasks.filter(due_date__gte=date_from)
    if date_to:
        tasks = tasks.filter(due_date__lte=date_to)

    tasks = tasks.distinct().order_by("-due_date")

    # Calculate advanced analysis metrics
    from django.utils import timezone
    from collections import defaultdict
    today = timezone.localdate()

    table_stats = defaultdict(lambda: {
        "name": "",
        "total": 0,
        "completed": 0,
        "overdue": 0,
        "pending": 0,
    })
    
    employee_stats = defaultdict(lambda: {
        "name": "",
        "department": "",
        "total": 0,
        "completed": 0,
        "overdue": 0,
        "pending": 0,
    })

    # Pre-evaluate and optimize querysets to prevent redundant queries
    tasks = tasks.select_related("row__table", "row__table__department", "assigned_by").prefetch_related("assigned_to", "row__cells__column")

    for t in tasks:
        tbl = t.row.table
        table_id = tbl.id
        if not table_stats[table_id]["name"]:
            table_stats[table_id]["name"] = tbl.name
        
        is_completed = t.status in ["COMPLETED", "APPROVED"]
        is_overdue = (t.due_date < today) and not is_completed
        
        table_stats[table_id]["total"] += 1
        if is_completed:
            table_stats[table_id]["completed"] += 1
        else:
            table_stats[table_id]["pending"] += 1
        if is_overdue:
            table_stats[table_id]["overdue"] += 1

        for user in t.assigned_to.all():
            user_id = user.id
            if not employee_stats[user_id]["name"]:
                employee_stats[user_id]["name"] = user.full_name
                employee_stats[user_id]["department"] = user.department.name if user.department else "Global"
            
            employee_stats[user_id]["total"] += 1
            if is_completed:
                employee_stats[user_id]["completed"] += 1
            else:
                employee_stats[user_id]["pending"] += 1
            if is_overdue:
                employee_stats[user_id]["overdue"] += 1

    for tid, stats in table_stats.items():
        stats["rate"] = f"{round((stats['completed'] / stats['total'] * 100), 2)}%" if stats["total"] > 0 else "0.0%"

    for uid, stats in employee_stats.items():
        stats["rate"] = f"{round((stats['completed'] / stats['total'] * 100), 2)}%" if stats["total"] > 0 else "0.0%"

    # Export formats
    export_format = request.GET.get("format")
    if export_format == "csv":
        response = HttpResponse(content_type="text/csv")
        response["Content-Disposition"] = 'attachment; filename="tasks_report.csv"'
        writer = csv.writer(response)
        
        # 1. Table Analysis Section
        writer.writerow(["=== TABLE ANALYTICS ==="])
        writer.writerow(["Table Name", "Total Tasks", "Completed Tasks", "Overdue Tasks", "Pending Tasks", "Completion Rate"])
        for tid, stats in table_stats.items():
            writer.writerow([stats["name"], stats["total"], stats["completed"], stats["overdue"], stats["pending"], stats["rate"]])
        writer.writerow([])
        
        # 2. Employee Analysis Section
        writer.writerow(["=== EMPLOYEE ANALYTICS ==="])
        writer.writerow(["Employee Name", "Department", "Total Tasks", "Completed Tasks", "Overdue Tasks", "Pending Tasks", "Achievement Rate"])
        for uid, stats in employee_stats.items():
            writer.writerow([stats["name"], stats["department"], stats["total"], stats["completed"], stats["overdue"], stats["pending"], stats["rate"]])
        writer.writerow([])
        
        # 3. Detailed Tasks Section
        writer.writerow(["=== DETAILED TASK REPORT ==="])
        writer.writerow(["S_NO", "Task Name", "Table Name", "Due Date", "Priority", "Status", "Assigned To", "Assigned By", "Department"])
        for t in tasks:
            task_name = t.row.cells.filter(column__name="TASK_NAME").first()
            task_name_val = task_name.value if task_name else "Unnamed"
            assigned_to = ", ".join([u.full_name for u in t.assigned_to.all()])
            assigned_by = t.assigned_by.full_name if t.assigned_by else "System"
            dept_name = t.row.table.department.name if t.row.table.department else "Global"
            s_no_cell = t.row.cells.filter(column__name="S_NO").first()
            s_no_val = s_no_cell.value if s_no_cell else t.id

            writer.writerow([s_no_val, task_name_val, t.row.table.name, t.due_date, t.priority, t.status, assigned_to, assigned_by, dept_name])
        return response

    elif export_format == "excel":
        wb = openpyxl.Workbook()
        
        # Sheet 1: Table Analytics
        ws1 = wb.active
        ws1.title = "Table Analytics"
        ws1.append(["Table Name", "Total Tasks", "Completed Tasks", "Overdue Tasks", "Pending Tasks", "Completion Rate"])
        for tid, stats in table_stats.items():
            ws1.append([stats["name"], stats["total"], stats["completed"], stats["overdue"], stats["pending"], stats["rate"]])
            
        # Sheet 2: Employee Analytics
        ws2 = wb.create_sheet(title="Employee Analytics")
        ws2.append(["Employee Name", "Department", "Total Tasks", "Completed Tasks", "Overdue Tasks", "Pending Tasks", "Achievement Rate"])
        for uid, stats in employee_stats.items():
            ws2.append([stats["name"], stats["department"], stats["total"], stats["completed"], stats["overdue"], stats["pending"], stats["rate"]])
            
        # Sheet 3: Detailed Tasks
        ws3 = wb.create_sheet(title="Detailed Tasks")
        ws3.append(["S_NO", "Task Name", "Table Name", "Due Date", "Priority", "Status", "Assigned To", "Assigned By", "Department"])
        for t in tasks:
            task_name = t.row.cells.filter(column__name="TASK_NAME").first()
            task_name_val = task_name.value if task_name else "Unnamed"
            assigned_to = ", ".join([u.full_name for u in t.assigned_to.all()])
            assigned_by = t.assigned_by.full_name if t.assigned_by else "System"
            dept_name = t.row.table.department.name if t.row.table.department else "Global"
            s_no_cell = t.row.cells.filter(column__name="S_NO").first()
            s_no_val = s_no_cell.value if s_no_cell else t.id

            ws3.append([s_no_val, task_name_val, t.row.table.name, str(t.due_date), t.priority, t.status, assigned_to, assigned_by, dept_name])

        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = 'attachment; filename="tasks_report.xlsx"'
        wb.save(response)
        return response

    elif export_format == "pdf":
        response = HttpResponse(content_type="application/pdf")
        response["Content-Disposition"] = 'attachment; filename="tasks_report.pdf"'
        
        doc = SimpleDocTemplate(response, pagesize=letter, leftMargin=36, rightMargin=36, topMargin=36, bottomMargin=36)
        story = []
        styles = getSampleStyleSheet()
        
        title_style = ParagraphStyle(
            name="TitleStyle",
            parent=styles["Heading1"],
            fontName="Helvetica-Bold",
            fontSize=20,
            leading=24,
            textColor=colors.HexColor("#1F2937"),
            spaceAfter=15
        )
        
        subtitle_style = ParagraphStyle(
            name="SubtitleStyle",
            parent=styles["Heading2"],
            fontName="Helvetica-Bold",
            fontSize=13,
            leading=16,
            textColor=colors.HexColor("#374151"),
            spaceBefore=12,
            spaceAfter=8
        )
        
        normal_style = ParagraphStyle(
            name="NormalStyle",
            parent=styles["Normal"],
            fontName="Helvetica",
            fontSize=8,
            leading=10
        )
        
        header_cell_style = ParagraphStyle(
            name="HeaderCellStyle",
            parent=styles["Normal"],
            fontName="Helvetica-Bold",
            fontSize=8,
            leading=10,
            textColor=colors.white
        )

        story.append(Paragraph("Flow-Force Workspace Advanced Analytics Report", title_style))
        story.append(Spacer(1, 10))
        
        # Section 1: Table Performance Analytics
        story.append(Paragraph("Table Performance Summary", subtitle_style))
        table_headers = ["Table Name", "Total Tasks", "Completed", "Overdue", "Pending", "Comp. Rate"]
        t_data = [[Paragraph(h, header_cell_style) for h in table_headers]]
        for tid, stats in table_stats.items():
            t_data.append([
                Paragraph(stats["name"], normal_style),
                Paragraph(str(stats["total"]), normal_style),
                Paragraph(str(stats["completed"]), normal_style),
                Paragraph(str(stats["overdue"]), normal_style),
                Paragraph(str(stats["pending"]), normal_style),
                Paragraph(stats["rate"], normal_style)
            ])
            
        t_style = TableStyle([
            ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#1F2937")),
            ("ALIGN", (0,0), (-1,-1), "LEFT"),
            ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
            ("BOTTOMPADDING", (0,0), (-1,0), 6),
            ("TOPPADDING", (0,0), (-1,0), 6),
            ("BOTTOMPADDING", (0,1), (-1,-1), 4),
            ("TOPPADDING", (0,1), (-1,-1), 4),
            ("BACKGROUND", (0,1), (-1,-1), colors.HexColor("#F9FAFB")),
            ("GRID", (0,0), (-1,-1), 0.5, colors.HexColor("#E5E7EB")),
        ])
        
        table_pdf = RLTable(t_data, colWidths=[200, 60, 70, 70, 70, 70])
        table_pdf.setStyle(t_style)
        story.append(table_pdf)
        story.append(Spacer(1, 15))
        
        # Section 2: Employee Performance Analytics
        story.append(Paragraph("Employee Achievement Summary", subtitle_style))
        emp_headers = ["Employee Name", "Department", "Total Tasks", "Completed", "Overdue", "Pending", "Ach. Rate"]
        emp_data = [[Paragraph(h, header_cell_style) for h in emp_headers]]
        for uid, stats in employee_stats.items():
            emp_data.append([
                Paragraph(stats["name"], normal_style),
                Paragraph(stats["department"], normal_style),
                Paragraph(str(stats["total"]), normal_style),
                Paragraph(str(stats["completed"]), normal_style),
                Paragraph(str(stats["overdue"]), normal_style),
                Paragraph(str(stats["pending"]), normal_style),
                Paragraph(stats["rate"], normal_style)
            ])
            
        emp_pdf = RLTable(emp_data, colWidths=[120, 110, 60, 60, 60, 60, 70])
        emp_pdf.setStyle(t_style)
        story.append(emp_pdf)
        story.append(Spacer(1, 15))
        
        # Section 3: Detailed Tasks List
        story.append(Paragraph("Detailed Task List", subtitle_style))
        task_headers = ["S_NO", "Task Name", "Table Name", "Due Date", "Priority", "Status", "Assigned To"]
        task_data = [[Paragraph(h, header_cell_style) for h in task_headers]]
        for t in tasks:
            task_name = t.row.cells.filter(column__name="TASK_NAME").first()
            task_name_val = task_name.value if task_name else "Unnamed"
            assigned_to = ", ".join([u.full_name for u in t.assigned_to.all()])
            s_no_cell = t.row.cells.filter(column__name="S_NO").first()
            s_no_val = s_no_cell.value if s_no_cell else t.id
            
            task_data.append([
                Paragraph(str(s_no_val), normal_style),
                Paragraph(task_name_val, normal_style),
                Paragraph(t.row.table.name, normal_style),
                Paragraph(str(t.due_date), normal_style),
                Paragraph(t.priority, normal_style),
                Paragraph(t.status, normal_style),
                Paragraph(assigned_to, normal_style)
            ])
            
        task_pdf = RLTable(task_data, colWidths=[40, 110, 90, 65, 55, 60, 120])
        task_pdf.setStyle(t_style)
        story.append(task_pdf)
        
        doc.build(story)
        return response

    context = {
        "employees": employees,
        "tasks": tasks,
    }
    return render(request, "tasks/reports.html", context)

@login_required
def task_detail_view(request, task_id):
    from django.contrib import messages
    task = get_object_or_404(Task, id=task_id)
    
    if request.method == "POST":
        content = request.POST.get("content")
        if content:
            # Create a new comment
            TaskComment.objects.create(
                task=task,
                author=request.user,
                content=content
            )
            # Create Activity Log
            ActivityLog.objects.create(
                task=task,
                action="Added Comment",
                user=request.user,
                details={}
            )
            # Send notification to other relevant users (creator, assignees, assigner)
            recipients = set()
            if task.assigned_by and task.assigned_by.id != request.user.id:
                recipients.add(task.assigned_by)
            for user in task.assigned_to.exclude(id=request.user.id):
                recipients.add(user)
            table_creator = task.row.table.created_by
            if table_creator and table_creator.id != request.user.id:
                recipients.add(table_creator)

            for recipient in recipients:
                Notification.objects.create(
                    user=recipient,
                    task=task,
                    title="New Comment on Flagged Task",
                    description=f"{request.user.full_name or request.user.email} replied to the task issue.",
                    type="COMMENT"
                )
            messages.success(request, "Comment posted successfully.")
            return redirect("tasks:task_detail", task_id=task.id)
            
    comments = task.comments.all().order_by("created_at")
    
    # Extract cells for details
    task_name = task.task_name
    table_name = task.table_name
    
    context = {
        "task": task,
        "task_name": task_name,
        "table_name": table_name,
        "comments": comments,
    }
    return render(request, "tasks/task_detail.html", context)

